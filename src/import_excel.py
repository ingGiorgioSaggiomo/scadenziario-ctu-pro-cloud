"""Import dati da file Excel del vecchio scadenziario.

Layout colonne:
  A = descrizione incarico       I = stato
  B = nomina                     J = giorni alla scadenza (IGNORATO, calcolato)
  C = giuramento                 K = note
  D = primo accesso              L = sospensione (data inizio)
  E = bozza                      M = ripresa (data fine)
  F = osservazioni CTP           N = giorni sospensione
  G = deposito
  H = udienza
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Optional, Union, IO

from openpyxl import load_workbook

from src.database import get_session
from src.models import Evento, Incarico, Sospensione


@dataclass
class ImportReport:
    imported: int = 0
    skipped: int = 0
    duplicates: int = 0
    date_errors: list[str] = field(default_factory=list)
    anomalies: list[str] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            "imported": self.imported,
            "skipped": self.skipped,
            "duplicates": self.duplicates,
            "date_errors": list(self.date_errors),
            "anomalies": list(self.anomalies),
        }


# ----- helpers di parsing ------------------------------------------------

_DATE_FORMATS = (
    "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d.%m.%Y",
    "%d/%m/%y", "%d-%m-%y",
)


@dataclass
class _DateResult:
    """Risultato del parsing di una cella data.

    Se valore parseable: parsed = la data, raw_text = None.
    Se testo non parseable: parsed = None, raw_text = stringa originale.
    Se cella vuota: parsed = None, raw_text = None.
    """
    parsed: Optional[date] = None
    raw_text: Optional[str] = None


def _parse_date_cell(value: Any, report: ImportReport, row: int, colname: str) -> _DateResult:
    if value is None:
        return _DateResult()
    if isinstance(value, datetime):
        return _DateResult(parsed=value.date())
    if isinstance(value, date):
        return _DateResult(parsed=value)
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            return _DateResult(parsed=from_excel(value).date())
        except Exception:
            report.date_errors.append(
                f"Riga {row} col {colname}: numero non interpretabile come data ({value!r})"
            )
            return _DateResult(raw_text=str(value))
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return _DateResult()
        for fmt in _DATE_FORMATS:
            try:
                return _DateResult(parsed=datetime.strptime(s, fmt).date())
            except ValueError:
                continue
        report.date_errors.append(
            f"Riga {row} col {colname}: testo non riconosciuto ({s!r})"
        )
        return _DateResult(raw_text=s)
    report.date_errors.append(
        f"Riga {row} col {colname}: tipo non gestito ({type(value).__name__})"
    )
    return _DateResult(raw_text=str(value))


def _normalize_stato(raw: Any) -> str:
    if raw is None:
        return "attivo"
    s = str(raw).strip().lower()
    if not s:
        return "attivo"
    if any(k in s for k in ("chius", "deposit", "complet", "conclu")):
        return "chiuso"
    if "sospes" in s:
        return "sospeso"
    return "attivo"


_RE_RG_PATTERNS = (
    re.compile(r"\b(?:RGE?|NR)\.?\s*(\d{1,6})\s*[./_-]\s*(\d{2,4})(?!\d)", re.IGNORECASE),
    re.compile(r"\bRESA\s+(\d{1,6})\s+(\d{2,4})(?!\d)", re.IGNORECASE),
    re.compile(r"\b(?:RGE?|NR)\.?\s*(\d{1,6})\s+(\d{2,4})(?!\d)", re.IGNORECASE),
    re.compile(r"\b(\d{1,6})\s*[./_-]\s*(\d{2,4})(?!\d)"),
)
_RE_TRIB = re.compile(
    r"(Tribunale di [\w'’ ]+|Procura della Repubblica di [\w'’ ]+|"
    r"Corte d['’]Appello di [\w'’ ]+|Giudice di Pace di [\w'’ ]+)",
    re.IGNORECASE,
)
def _parse_descrizione(desc: str) -> dict:
    out = {"tipo": "CTU", "numero_rg": None, "tribunale": None, "oggetto": desc}
    first_line = desc.splitlines()[0].strip()
    upper = first_line.upper()
    normalized = re.sub(r"[^A-Z]", "", upper)

    if "PROCURA" in upper:
        out["tipo"] = "Procura"
    elif "ORDINEAFARE" in normalized:
        out["tipo"] = "Ordine a fare"
    elif "INCARICOTECNICO" in normalized:
        out["tipo"] = "Incarico tecnico"
    elif "RESA" in normalized:
        out["tipo"] = "RESA"
    elif "CTU" in upper:
        out["tipo"] = "CTU"

    for pattern in _RE_RG_PATTERNS:
        m = pattern.search(first_line)
        if m:
            out["numero_rg"] = f"{m.group(1)}/{m.group(2)}"
            break
    m = _RE_TRIB.search(desc)
    if m:
        out["tribunale"] = m.group(1).strip()
    return out


def _is_duplicate(session, oggetto: str, data_nomina: Optional[date],
                  data_iniz: Optional[date]) -> bool:
    """Controlla se esiste già un incarico con stessa terna oggetto/nomina/inizio."""
    q = session.query(Incarico).filter(
        Incarico.oggetto == oggetto,
        Incarico.data_conferimento == data_nomina,
    )
    if data_iniz is None:
        q = q.filter(Incarico.data_inizio_operazioni.is_(None))
    else:
        q = q.filter(Incarico.data_inizio_operazioni == data_iniz)
    return session.query(q.exists()).scalar()


# ----- importazione principale -------------------------------------------

def importa_excel(
    file: Union[str, IO[bytes]],
    sheet_name: Optional[str] = None,
    start_row: int = 2,
) -> dict:
    """Importa dati dallo scadenziario Excel.

    Args:
        file: percorso o oggetto file-like (es. UploadedFile di Streamlit).
        sheet_name: nome del foglio; se None usa il primo.
        start_row: riga iniziale dei dati (1-based).

    Returns:
        report come dict (vedi ImportReport.as_dict).
    """
    report = ImportReport()
    wb = load_workbook(file, data_only=True, read_only=False)
    ws = wb[sheet_name] if sheet_name else wb.active

    if ws.merged_cells.ranges:
        report.anomalies.append(
            f"Trovate {len(ws.merged_cells.ranges)} celle unite: "
            f"i valori vengono letti solo dalla cella di ancoraggio."
        )

    session = get_session()
    try:
        for r in range(start_row, ws.max_row + 1):
            descr_raw = ws.cell(row=r, column=1).value
            if descr_raw is None or (isinstance(descr_raw, str) and not descr_raw.strip()):
                report.skipped += 1
                continue

            descr = str(descr_raw).strip()
            parsed = _parse_descrizione(descr)

            d_nomina = _parse_date_cell(ws.cell(row=r, column=2).value, report, r, "B/nomina")
            d_giura = _parse_date_cell(ws.cell(row=r, column=3).value, report, r, "C/giuramento")
            d_iniz = _parse_date_cell(ws.cell(row=r, column=4).value, report, r, "D/inizio op.")
            d_bozza = _parse_date_cell(ws.cell(row=r, column=5).value, report, r, "E/bozza")
            d_oss = _parse_date_cell(ws.cell(row=r, column=6).value, report, r, "F/osservazioni")
            d_dep = _parse_date_cell(ws.cell(row=r, column=7).value, report, r, "G/deposito")
            d_ud = _parse_date_cell(ws.cell(row=r, column=8).value, report, r, "H/udienza")

            stato_raw = ws.cell(row=r, column=9).value
            stato = _normalize_stato(stato_raw)

            # Colonna J = giorni alla scadenza, IGNORATA per design

            note_raw = ws.cell(row=r, column=11).value
            note = str(note_raw).strip() if note_raw is not None and str(note_raw).strip() else None

            d_sosp = _parse_date_cell(ws.cell(row=r, column=12).value, report, r, "L/sospensione")
            d_ripr = _parse_date_cell(ws.cell(row=r, column=13).value, report, r, "M/ripresa")
            gg_sosp = ws.cell(row=r, column=14).value

            data_nomina_val = d_nomina.parsed
            if data_nomina_val is None:
                fallback = d_giura.parsed or d_iniz.parsed
                if fallback:
                    data_nomina_val = fallback
                    report.anomalies.append(
                        f"Riga {r}: data nomina mancante, usata in fallback {fallback.isoformat()}"
                    )
                else:
                    data_nomina_val = date.today()
                    report.anomalies.append(
                        f"Riga {r}: nessuna data utile, impostata data odierna come nomina"
                    )

            # Anti-duplicazione
            if _is_duplicate(session, descr, data_nomina_val, d_iniz.parsed):
                report.duplicates += 1
                report.anomalies.append(
                    f"Riga {r}: duplicato (oggetto+nomina+inizio op. già presenti), saltato"
                )
                continue

            if not parsed["numero_rg"]:
                parsed["numero_rg"] = f"IMPORT-{r}"
                report.anomalies.append(
                    f"Riga {r}: numero RG non rilevato dalla descrizione, generato {parsed['numero_rg']}"
                )

            tribunale = parsed["tribunale"] or "(da definire)"
            if parsed["tribunale"] is None:
                report.anomalies.append(f"Riga {r}: ufficio non rilevato, impostato '(da definire)'")

            inc = Incarico(
                tipo=parsed["tipo"],
                numero_rg=parsed["numero_rg"],
                tribunale=tribunale,
                oggetto=descr,
                data_conferimento=data_nomina_val,
                data_giuramento=d_giura.parsed,
                data_inizio_operazioni=d_iniz.parsed,
                stato=stato,
                priorita="media",
                origine_dato="import_excel",
                note=note,
            )

            # Eventi standard E-H: parseable -> evento normale; testo -> evento senza data con nota
            for d_res, tipo_ev, colname in (
                (d_bozza, "bozza", "E"),
                (d_oss, "osservazioni", "F"),
                (d_dep, "deposito", "G"),
                (d_ud, "udienza", "H"),
            ):
                if d_res.parsed:
                    inc.eventi.append(Evento(
                        tipo=tipo_ev,
                        data=d_res.parsed,
                        descrizione="Importato da Excel",
                    ))
                elif d_res.raw_text:
                    inc.eventi.append(Evento(
                        tipo=tipo_ev,
                        data=None,
                        descrizione=f"Testo non interpretato in colonna {colname}: {d_res.raw_text!r}",
                    ))

            # Testo non parseable in colonne B/C/D/L/M -> nota evento
            for d_res, colname in (
                (d_nomina, "B"), (d_giura, "C"), (d_iniz, "D"),
                (d_sosp, "L"), (d_ripr, "M"),
            ):
                if d_res.raw_text and not d_res.parsed:
                    inc.eventi.append(Evento(
                        tipo="nota",
                        data=None,
                        descrizione=f"Testo non interpretato in colonna {colname}: {d_res.raw_text!r}",
                    ))

            if d_sosp.parsed:
                motivo = "Importata da Excel"
                if gg_sosp not in (None, ""):
                    motivo += f" — giorni dichiarati: {gg_sosp}"
                # Le sospensioni importate hanno valore annotativo: non incidono
                # automaticamente sulle scadenze salvo conferma manuale.
                inc.sospensioni.append(Sospensione(
                    data_inizio=d_sosp.parsed,
                    data_fine=d_ripr.parsed,
                    motivo=motivo,
                    incide_su_scadenze=False,
                ))

            session.add(inc)
            report.imported += 1

        session.commit()
    except Exception as exc:
        session.rollback()
        report.anomalies.append(f"Errore durante l'importazione: {exc!r}")
        raise
    finally:
        session.close()

    return report.as_dict()
