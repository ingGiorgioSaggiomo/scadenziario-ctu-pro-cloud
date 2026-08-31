"""Export verso Excel."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.models import Documento, Evento, Incarico, ModificaChat, Pagamento, Sospensione, StoricoTermine, Termine
from src.pagamenti import calcola_totale_fattura, importo_dovuto_pagamento
from src.utils import calcola_scadenza_termine, stato_evento


DATE_FORMAT = "DD/MM/YYYY"


def _label_incarico(incarico: Incarico) -> str:
    return f"{incarico.tipo} {incarico.numero_rg} - {incarico.tribunale}"


def _append_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    if workbook.active.title == "Sheet" and workbook.active.max_row == 1 and workbook.active.max_column == 1:
        worksheet = workbook.active
        worksheet.title = title
    else:
        worksheet = workbook.create_sheet(title)

    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if hasattr(cell.value, "year") and hasattr(cell.value, "month") and hasattr(cell.value, "day"):
                cell.number_format = DATE_FORMAT

    for col_idx, column_cells in enumerate(worksheet.columns, start=1):
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 45)


def genera_excel_export(session) -> bytes:
    """Genera un file Excel con i dati correnti dello scadenziario."""
    incarichi = session.query(Incarico).order_by(Incarico.id).all()
    termini = session.query(Termine).order_by(Termine.incarico_id, Termine.id).all()
    eventi = session.query(Evento).order_by(Evento.incarico_id, Evento.id).all()
    sospensioni = session.query(Sospensione).order_by(Sospensione.incarico_id, Sospensione.id).all()
    documenti = session.query(Documento).order_by(Documento.incarico_id, Documento.id).all()
    pagamenti = session.query(Pagamento).order_by(Pagamento.incarico_id, Pagamento.id).all()
    modifiche_chat = session.query(ModificaChat).order_by(ModificaChat.id).all()
    storico_termini = session.query(StoricoTermine).order_by(StoricoTermine.id).all()

    incarichi_by_id = {inc.id: inc for inc in incarichi}

    workbook = Workbook()

    _append_sheet(
        workbook,
        "Incarichi",
        [
            "ID", "Tipo", "Numero procedura", "Ufficio", "Giudice/PM",
            "Parti", "Oggetto", "Data nomina", "Data giuramento",
            "Data inizio operazioni", "Data invio bozza",
            "Data ricezione osservazioni", "Stato", "Priorita",
            "Origine", "Note",
        ],
        [
            [
                inc.id, inc.tipo, inc.numero_rg, inc.tribunale, inc.giudice,
                inc.parti, inc.oggetto, inc.data_conferimento, inc.data_giuramento,
                inc.data_inizio_operazioni, inc.data_invio_bozza,
                inc.data_ricezione_osservazioni, inc.stato, inc.priorita,
                inc.origine_dato, inc.note,
            ]
            for inc in incarichi
        ],
    )

    _append_sheet(
        workbook,
        "Termini",
        [
            "ID", "Incarico ID", "Incarico", "Tipo termine", "Giorni",
            "Decorrenza", "Data manuale", "Scadenza calcolata",
            "Attivo", "Completato", "Prorogato", "Note",
        ],
        [
            [
                term.id, term.incarico_id, _label_incarico(incarichi_by_id[term.incarico_id]),
                term.tipo_termine, term.giorni, term.decorrenza, term.data_manual,
                calcola_scadenza_termine(incarichi_by_id[term.incarico_id], term),
                term.attivo, term.completato, term.prorogato, term.note,
            ]
            for term in termini
            if term.incarico_id in incarichi_by_id
        ],
    )

    _append_sheet(
        workbook,
        "Eventi",
        [
            "ID", "Incarico ID", "Incarico", "Tipo", "Data", "Ora",
            "Luogo", "Descrizione", "Stato", "Termine collegato ID",
        ],
        [
            [
                evento.id, evento.incarico_id, _label_incarico(incarichi_by_id[evento.incarico_id]),
                evento.tipo, evento.data, evento.ora, evento.luogo,
                evento.descrizione, stato_evento(evento), evento.termine_id,
            ]
            for evento in eventi
            if evento.incarico_id in incarichi_by_id
        ],
    )

    _append_sheet(
        workbook,
        "Sospensioni",
        [
            "ID", "Incarico ID", "Incarico", "Data sospensione",
            "Data ripresa", "Incide su scadenze", "Motivo",
        ],
        [
            [
                sosp.id, sosp.incarico_id, _label_incarico(incarichi_by_id[sosp.incarico_id]),
                sosp.data_inizio, sosp.data_fine, sosp.incide_su_scadenze, sosp.motivo,
            ]
            for sosp in sospensioni
            if sosp.incarico_id in incarichi_by_id
        ],
    )

    _append_sheet(
        workbook,
        "Documenti",
        [
            "ID", "Incarico ID", "Incarico", "Nome", "Percorso",
            "Tipo", "Data documento", "Note",
        ],
        [
            [
                doc.id, doc.incarico_id, _label_incarico(incarichi_by_id[doc.incarico_id]),
                doc.nome, doc.percorso, doc.tipo, doc.data_documento, doc.note,
            ]
            for doc in documenti
            if doc.incarico_id in incarichi_by_id
        ],
    )

    _append_sheet(
        workbook,
        "Pagamenti",
        [
            "ID", "Incarico ID", "Incarico", "Tipo", "Descrizione",
            "Imponibile", "Spese liquidate", "Cassa 4%", "Marca da bollo",
            "Importo dovuto da incassare", "Importo ricevuto",
            "Data riferimento", "Data pagamento", "Pagatore", "Note",
        ],
        [
            [
                pagamento.id, pagamento.incarico_id,
                _label_incarico(incarichi_by_id[pagamento.incarico_id]),
                pagamento.tipo, pagamento.descrizione,
                float(pagamento.imponibile or 0),
                float(pagamento.spese or 0),
                float(calcola_totale_fattura(pagamento.imponibile, pagamento.spese)["cassa"]),
                float(calcola_totale_fattura(pagamento.imponibile, pagamento.spese)["bollo"]),
                float(importo_dovuto_pagamento(pagamento)),
                float(pagamento.importo_ricevuto or 0),
                pagamento.data_riferimento, pagamento.data_pagamento,
                pagamento.pagatore, pagamento.note,
            ]
            for pagamento in pagamenti
            if pagamento.incarico_id in incarichi_by_id
        ],
    )

    _append_sheet(
        workbook,
        "Registro modifiche",
        [
            "ID", "Incarico ID", "Incarico", "Data", "Azione",
            "Richiesta", "Prima", "Dopo", "Note",
        ],
        [
            [
                modifica.id, modifica.incarico_id,
                _label_incarico(incarichi_by_id[modifica.incarico_id])
                if modifica.incarico_id in incarichi_by_id else "",
                modifica.created_at, modifica.azione, modifica.richiesta,
                modifica.prima, modifica.dopo, modifica.note,
            ]
            for modifica in modifiche_chat
        ],
    )

    _append_sheet(
        workbook,
        "Storico termini",
        [
            "ID", "Incarico ID", "Incarico", "Termine ID", "Data modifica",
            "Azione", "Motivo", "Tipo", "Giorni", "Decorrenza", "Data manuale",
            "Scadenza", "Attivo", "Completato", "Prorogato", "Note",
        ],
        [
            [
                voce.id, voce.incarico_id,
                _label_incarico(incarichi_by_id[voce.incarico_id])
                if voce.incarico_id in incarichi_by_id else "",
                voce.termine_id, voce.modificato_il, voce.azione, voce.motivo,
                voce.tipo_termine, voce.giorni, voce.decorrenza, voce.data_manual,
                voce.data_scadenza, voce.attivo, voce.completato, voce.prorogato, voce.note,
            ]
            for voce in storico_termini
        ],
    )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
