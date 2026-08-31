"""Test per export Excel."""

from datetime import date
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from src.export_tools import genera_excel_export
from src.models import Base, Documento, Evento, Incarico, ModificaChat, Pagamento, Sospensione, StoricoTermine, Termine


def _session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine)()


def _as_date(value):
    return value.date() if hasattr(value, "date") else value


def test_export_excel_contiene_fogli_e_dati():
    session = _session()
    inc = Incarico(
        tipo="CTU",
        numero_rg="57/2024",
        tribunale="Tribunale di Napoli",
        data_conferimento=date(2026, 1, 1),
        data_inizio_operazioni=date(2026, 1, 10),
        stato="attivo",
        priorita="media",
        origine_dato="manuale",
    )
    inc.termini = [
        Termine(
            tipo_termine="bozza",
            giorni=5,
            decorrenza="data_inizio_operazioni",
            tipo_computo="naturali",
            attivo=True,
        )
    ]
    inc.eventi = [Evento(tipo="udienza", data=date(2026, 2, 1), completato=True)]
    inc.sospensioni = [Sospensione(data_inizio=date(2026, 3, 1), incide_su_scadenze=False)]
    inc.documenti = [Documento(nome="relazione.pdf", tipo="perizia")]
    inc.pagamenti = [Pagamento(tipo="acconto", imponibile=1000, spese=100, importo_dovuto=1142, importo_ricevuto=500)]
    session.add(inc)
    session.commit()
    session.add(
        ModificaChat(
            incarico_id=inc.id,
            azione="aggiungi_pagamento",
            richiesta="aggiungi acconto",
            prima="prima",
            dopo="dopo",
        )
    )
    session.add(
        StoricoTermine(
            incarico_id=inc.id,
            termine_id=inc.termini[0].id,
            azione="creato",
            tipo_termine="bozza",
            giorni=5,
            decorrenza="data_inizio_operazioni",
            data_scadenza=date(2026, 1, 15),
            attivo=True,
            completato=False,
            prorogato=False,
        )
    )
    session.commit()

    data = genera_excel_export(session)
    workbook = load_workbook(BytesIO(data), data_only=True)

    assert workbook.sheetnames == [
        "Incarichi", "Termini", "Eventi", "Sospensioni", "Documenti", "Pagamenti", "Registro modifiche",
        "Storico termini",
    ]
    assert workbook["Incarichi"]["C2"].value == "57/2024"
    incarichi_headers = [cell.value for cell in workbook["Incarichi"][1]]
    assert "Origine iniziale" in incarichi_headers
    assert "Origine" not in incarichi_headers
    assert workbook["Eventi"]["I2"].value == "completato"
    assert workbook["Pagamenti"]["D2"].value == "acconto"
    assert workbook["Pagamenti"]["F2"].value == 1000
    assert workbook["Pagamenti"]["G2"].value == 100
    assert workbook["Pagamenti"]["H2"].value == 40
    assert workbook["Pagamenti"]["I2"].value == 2
    assert workbook["Pagamenti"]["J2"].value == 1142
    assert workbook["Registro modifiche"]["E2"].value == "aggiungi_pagamento"
    assert workbook["Storico termini"]["F2"].value == "creato"

    term_headers = [cell.value for cell in workbook["Termini"][1]]
    scadenza_col = term_headers.index("Scadenza calcolata") + 1
    assert _as_date(workbook["Termini"].cell(2, scadenza_col).value) == date(2026, 1, 15)
    session.close()


def test_export_excel_senza_dati_crea_solo_intestazioni():
    session = _session()
    data = genera_excel_export(session)
    workbook = load_workbook(BytesIO(data), data_only=True)

    assert workbook["Incarichi"].max_row == 1
    assert workbook["Termini"].max_row == 1
    assert workbook["Eventi"].max_row == 1
    assert workbook["Pagamenti"].max_row == 1
    assert workbook["Registro modifiche"].max_row == 1
    assert workbook["Storico termini"].max_row == 1
    session.close()
